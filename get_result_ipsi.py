import os
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import chardet
import streamlit as st

# Streamlit 앱 제목
st.title('경쟁률 파일 업로드 및 처리')

# 파일을 저장할 경로
UPLOAD_FOLDER = './uploads/'
OUTPUT_FOLDER = './경쟁률모음/'

# 파일 저장 디렉토리가 없을 경우 생성
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# 결과 파일 저장할 '경쟁률모음' 폴더 생성
if not os.path.exists(OUTPUT_FOLDER):
    os.makedirs(OUTPUT_FOLDER)

# 파일 업로드 위젯
uploaded_file = st.file_uploader("경쟁률url.xlsx 파일을 업로드하세요", type="xlsx")

# 파일이 업로드되었을 때 처리
if uploaded_file is not None:
    # 업로드된 파일을 저장할 경로
    file_path = os.path.join(UPLOAD_FOLDER, uploaded_file.name)
    
    # 업로드된 파일을 파일로 저장
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    
    st.success(f"파일 '{uploaded_file.name}' 업로드 및 저장 완료!")

    # 파일을 pandas로 읽기
    df = pd.read_excel(file_path)

    # 데이터프레임 미리 보기
    st.write("업로드된 파일의 내용:")
    st.dataframe(df)

    # 경쟁률 URL 처리 함수 (파일이 있으면 시트를 추가하고, 없으면 새로 생성)
    def scrape_and_save_to_excel(university_name, url):
        response = requests.get(url, timeout=10)
        encoding = chardet.detect(response.content)['encoding']
        response.encoding = encoding

        soup = BeautifulSoup(response.text, 'html.parser')

        # 파일 경로
        file_name = f"{OUTPUT_FOLDER}{university_name}.xlsx"
        
        # 이미 파일이 존재하는 경우: 기존 파일에 시트 추가
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.create_sheet(title=f"Web Data {len(wb.sheetnames)}")
            st.write(f"{university_name}에 새 시트를 추가합니다.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Web Data"
            st.write(f"{university_name} 파일을 새로 생성합니다.")

        # 데이터 처리 및 엑셀에 기록
        all_elements = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'p', 'table'])

        for element in all_elements:
            if element.name == 'table':
                headers = [header.text.strip() for header in element.find_all('th')]
                if headers:
                    ws.append(headers)

                rows = element.find_all('tr')
                for row in rows:
                    cols = [col.text.strip() for col in row.find_all('td')]
                    if cols:
                        ws.append(cols)

                ws.append([])

            else:
                text = element.get_text(strip=True)
                if text:
                    ws.append([text])
                    ws.append([])

        # 파일 저장
        wb.save(file_name)
        st.write(f"{university_name}의 데이터를 '{file_name}'에 저장 완료!")

    # 업로드된 파일에 있는 각 대학의 URL 처리
    skipped_universities = []
    
    for index, row in df.iterrows():
        university_name = row['대학명']
        url = row['url']
        
        if pd.isna(url) or url.strip() == "":
            st.write(f"{university_name}의 URL이 없어서 건너뜁니다.")
            continue
        
        try:
            scrape_and_save_to_excel(university_name, url)
        except requests.exceptions.RequestException as e:
            skipped_universities.append(university_name)
            st.write(f"오류 발생: {university_name}의 URL({url})에서 오류가 발생했습니다.")

    # 패스된 대학 출력
    if skipped_universities:
        st.write("\n오류가 발생하여 패스된 대학들:")
        for uni in skipped_universities:
            st.write(f"- {uni}")
    else:
        st.write("\n모든 대학의 URL에서 정상적으로 데이터를 가져왔습니다.")
